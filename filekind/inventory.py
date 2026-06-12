from __future__ import annotations

import re
from pathlib import Path

from openpyxl import load_workbook

from filekind.models import ProjectDef

PROJECT_CODE_IN_TEXT = re.compile(r"(?i)\d{6}--[A-Z0-9]+")
INVENTORY_SERIAL = re.compile(r"(?i)^20\d{4}$")

CODE_HEADERS = ("项目编号", "编号", "项目代号", "代号", "项目代码", "code")
NAME_HEADERS = ("项目名称", "系统名称", "产品名称", "名称", "系统", "name", "title")
YEAR_HEADERS = ("年份", "year")
SOLUTION_HEADERS = ("方案名", "方案名称", "solution")
BOARD_HEADERS = ("版型名称", "版型", "板型", "board")


class InventoryError(ValueError):
    pass


def _normalize_header(value: object) -> str:
    return str(value or "").strip().casefold()


def _cell_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _slug_id(code: str, name: str, index: int) -> str:
    base = code or name or f"project-{index}"
    slug = re.sub(r"[^\w\-]+", "-", base.casefold()).strip("-")
    return slug[:80] or f"project-{index}"


def _extract_legacy_code(text: str) -> str:
    match = PROJECT_CODE_IN_TEXT.search(text)
    return match.group(0).upper() if match else ""


def _normalize_inventory_code(raw: str) -> tuple[str, str]:
    """Return (inventory_code, legacy_code_for_codes_list)."""
    text = raw.strip()
    if not text:
        return "", ""
    legacy = _extract_legacy_code(text)
    if legacy:
        return legacy, legacy
    serial = text.upper()
    if INVENTORY_SERIAL.fullmatch(serial):
        return serial, serial
    return serial, serial


def _header_alias_match(header: str, aliases: tuple[str, ...]) -> bool:
    return any(
        header == alias.casefold() or header.endswith(alias.casefold()) for alias in aliases
    )


def _find_header_map(row_values: list[str]) -> dict[str, int] | None:
    normalized = [_normalize_header(v) for v in row_values]
    indices: dict[str, int] = {}
    for idx, header in enumerate(normalized):
        if not header:
            continue
        if "code" not in indices and _header_alias_match(header, CODE_HEADERS):
            indices["code"] = idx
        if "name" not in indices and _header_alias_match(header, NAME_HEADERS):
            indices["name"] = idx
        if "year" not in indices and _header_alias_match(header, YEAR_HEADERS):
            indices["year"] = idx
        if "solution" not in indices and _header_alias_match(header, SOLUTION_HEADERS):
            indices["solution"] = idx
        if "board" not in indices and _header_alias_match(header, BOARD_HEADERS):
            indices["board"] = idx
    if "code" not in indices and "name" not in indices:
        return None
    return indices


def _locate_header_row(rows: list[tuple[object, ...]]) -> tuple[int, dict[str, int] | None]:
    for index, row in enumerate(rows[:20]):
        header_map = _find_header_map([_cell_text(c) for c in row])
        if header_map and "name" in header_map:
            return index, header_map
    return 0, _find_header_map([_cell_text(c) for c in rows[0]]) if rows else None


def _cell_at(row: tuple[object, ...], index: int) -> str:
    if index < 0 or index >= len(row):
        return ""
    return _cell_text(row[index])


def _projects_from_sheet_rows(rows: list[tuple[object, ...]]) -> list[ProjectDef]:
    if not rows:
        return []

    header_index, header_map = _locate_header_row(rows)
    data_rows = rows[header_index + 1 :] if header_map else rows

    projects: list[ProjectDef] = []
    seen: set[str] = set()

    for row in data_rows:
        cells = [_cell_text(c) for c in row if _cell_text(c)]
        if not cells:
            continue

        raw_code = ""
        name = ""
        year = ""
        solution_name = ""
        board_type = ""
        if header_map:
            if "code" in header_map:
                raw_code = _cell_at(row, header_map["code"])
            if "name" in header_map:
                name = _cell_at(row, header_map["name"])
            if "year" in header_map:
                year = _cell_at(row, header_map["year"])
            if "solution" in header_map:
                solution_name = _cell_at(row, header_map["solution"])
            if "board" in header_map:
                board_type = _cell_at(row, header_map["board"])

        row_text = " ".join(cells)
        inventory_code, code_for_list = _normalize_inventory_code(raw_code)
        if not inventory_code:
            inventory_code, code_for_list = _normalize_inventory_code(_extract_legacy_code(row_text))
        if not name:
            name = max(cells, key=len)
        if not inventory_code and name:
            inventory_code, code_for_list = _normalize_inventory_code(_extract_legacy_code(name))

        if not inventory_code and not name:
            continue

        dedupe_key = (inventory_code or name).casefold()
        if dedupe_key in seen:
            continue
        seen.add(dedupe_key)

        codes: list[str] = []
        if code_for_list:
            codes.append(code_for_list)

        aliases: list[str] = []
        if solution_name:
            aliases.append(solution_name)
        if board_type and board_type.casefold() not in name.casefold():
            aliases.append(board_type)
        if inventory_code and inventory_code.casefold() not in name.casefold():
            aliases.append(inventory_code)

        projects.append(
            ProjectDef(
                id=_slug_id(inventory_code or code_for_list, name, len(projects) + 1),
                name=name or inventory_code,
                aliases=list(dict.fromkeys(aliases)),
                codes=codes,
                description=row_text[:500],
                inventory_code=inventory_code,
                solution_name=solution_name,
                board_type=board_type,
                year=year,
            )
        )

    return projects


def looks_like_inventory_workbook(path: Path) -> bool:
    """True when the workbook has a recognizable project-list header row."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception:
        return False
    try:
        for sheet in wb.worksheets:
            rows: list[tuple[object, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows.append(row)
                if len(rows) >= 20:
                    break
            for row in rows[:20]:
                header_map = _find_header_map([_cell_text(c) for c in row])
                if header_map and "name" in header_map:
                    return True
        return False
    finally:
        wb.close()


def load_projects_from_inventory(path: Path) -> list[ProjectDef]:
    """Parse project list from an Excel inventory workbook."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
    except Exception as exc:
        raise InventoryError(f"无法读取项目清单 Excel: {path.name} ({exc})") from exc

    projects: list[ProjectDef] = []
    seen_ids: set[str] = set()
    try:
        for sheet in wb.worksheets:
            rows: list[tuple[object, ...]] = []
            for row in sheet.iter_rows(values_only=True):
                if any(cell is not None and str(cell).strip() for cell in row):
                    rows.append(row)
                if len(rows) >= 2000:
                    break
            for project in _projects_from_sheet_rows(rows):
                if project.id in seen_ids:
                    project.id = f"{project.id}-{len(seen_ids)}"
                seen_ids.add(project.id)
                projects.append(project)
    finally:
        wb.close()

    if not projects:
        raise InventoryError(
            f"项目清单 Excel 中未解析到任何项目: {path.name}\n"
            "请确认表格含有「项目编号/项目名称」等列，或每行包含如 202432--CSP311 的项目代号。"
        )
    return projects
