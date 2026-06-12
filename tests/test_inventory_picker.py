from pathlib import Path

import pytest
from openpyxl import Workbook

from filekind.inventory_picker import (
    discover_inventory_candidates,
    is_temp_excel,
    load_last_inventory,
    resolve_inventory_for_run,
    save_last_inventory,
    try_count_inventory_projects,
)
from filekind.models import AppConfig


def _write_inventory(path: Path, *, name: str = "清单A") -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["项目编号", "项目名称"])
    ws.append(["202432--CSP311", name])
    wb.save(path)


def _write_board_inventory(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["年份", "项目编号", "项目名称", "方案名", "版型名称"])
    ws.append(["2023", "202301", "CV950D4系统", "CV950D4", "CV950D4-B42"])
    wb.save(path)


def test_is_temp_excel() -> None:
    assert is_temp_excel(Path("~$清单.xlsx"))
    assert not is_temp_excel(Path("清单.xlsx"))


def test_discover_skips_non_inventory_xlsx(tmp_path: Path) -> None:
    source = tmp_path / "待整理"
    source.mkdir()
    inv = source / "国高资料清单.xlsx"
    noise = source / "测试报告.xlsx"
    _write_inventory(inv)
    wb = Workbook()
    wb.active.append(["日期", "金额"])
    wb.active.append(["2024-01-01", 100])
    wb.save(noise)

    config_path = tmp_path / "projects.yaml"
    config_path.write_text("paths: {}\n", encoding="utf-8")

    candidates = discover_inventory_candidates(config_path, source)
    assert len(candidates) == 1
    assert candidates[0][0] == inv.resolve()


def test_auto_pick_single_inventory(tmp_path: Path) -> None:
    source = tmp_path / "待整理"
    source.mkdir()
    inv = source / "list.xlsx"
    _write_inventory(inv)
    config_path = tmp_path / "projects.yaml"
    config_path.write_text("paths: {}\n", encoding="utf-8")
    config = AppConfig()

    picked = resolve_inventory_for_run(
        config_path,
        config,
        source,
        interactive=False,
    )
    assert picked == inv.resolve()
    assert load_last_inventory(config_path) == inv.resolve()


def test_explicit_inventory_path(tmp_path: Path) -> None:
    inv = tmp_path / "manual.xlsx"
    _write_board_inventory(inv)
    config_path = tmp_path / "projects.yaml"
    config_path.touch()
    config = AppConfig()
    source = tmp_path / "待整理"
    source.mkdir()

    picked = resolve_inventory_for_run(
        config_path,
        config,
        source,
        explicit=inv,
        interactive=False,
    )
    assert picked == inv.resolve()


def test_multiple_inventories_require_interactive(tmp_path: Path) -> None:
    source = tmp_path / "待整理"
    source.mkdir()
    _write_inventory(source / "a.xlsx", name="项目A")
    _write_inventory(source / "b.xlsx", name="项目B")
    config_path = tmp_path / "projects.yaml"
    config_path.write_text("paths: {}\n", encoding="utf-8")
    config = AppConfig()

    with pytest.raises(Exception, match="多个项目清单"):
        resolve_inventory_for_run(
            config_path,
            config,
            source,
            interactive=False,
        )


def test_save_and_load_last_inventory(tmp_path: Path) -> None:
    config_path = tmp_path / "projects.yaml"
    config_path.touch()
    inv = tmp_path / "list.xlsx"
    _write_inventory(inv)
    save_last_inventory(config_path, inv)
    assert load_last_inventory(config_path) == inv.resolve()


def test_try_count_inventory_projects(tmp_path: Path) -> None:
    inv = tmp_path / "list.xlsx"
    _write_inventory(inv)
    assert try_count_inventory_projects(inv) == 1
    assert try_count_inventory_projects(tmp_path / "missing.xlsx") is None
