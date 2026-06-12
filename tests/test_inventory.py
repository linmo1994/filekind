from pathlib import Path

import pytest
from openpyxl import Workbook

from filekind.classify.llm import _compact_candidates
from filekind.classify.rules import apply_rules, collect_signals
from filekind.config import resolve_inventory_path
from filekind.inventory import InventoryError, load_projects_from_inventory
from filekind.models import AppConfig, FileRecord, PathsConfig, ProjectDef, RuntimeConfig
from filekind.scan.scanner import scan_directory


def _write_inventory(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["项目编号", "项目名称"])
    ws.append(["202432--CSP311", "基于Android T高级智能会议平板系统"])
    ws.append(["202414--CV6683", "基于Android R南美智能数字电视系统"])
    wb.save(path)


def _write_board_inventory(path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["研发项目可行性报告清单"])
    ws.append(["年份", "项目编号", "项目名称", "备注", "方案名", "版型名称"])
    ws.append(
        [
            "2023",
            "202301",
            "CV950D4基于Android R的欧规全制式智能数字电视系统软件",
            "杨灏",
            "CV950D4",
            "CV950D4-B42",
        ]
    )
    ws.append(
        [
            "2024",
            "202401",
            "CV950D4基于Android 13的欧洲全制式智能数字电视系统软件",
            "杨灏",
            "CV950D4",
            "CV950D4-B42",
        ]
    )
    wb.save(path)


def test_load_projects_from_inventory(tmp_path: Path) -> None:
    xlsx = tmp_path / "国高资料清单2026 版型清单.xlsx"
    _write_inventory(xlsx)
    projects = load_projects_from_inventory(xlsx)
    assert len(projects) == 2
    assert projects[0].codes == ["202432--CSP311"]
    assert "会议平板" in projects[0].name


def test_load_board_inventory_with_title_row(tmp_path: Path) -> None:
    xlsx = tmp_path / "board.xlsx"
    _write_board_inventory(xlsx)
    projects = load_projects_from_inventory(xlsx)
    assert len(projects) == 2
    first = projects[0]
    assert first.inventory_code == "202301"
    assert first.codes == ["202301"]
    assert first.solution_name == "CV950D4"
    assert first.board_type == "CV950D4-B42"
    assert first.year == "2023"


def test_rule_matches_cv950_spec_by_inventory_code(tmp_path: Path) -> None:
    xlsx = tmp_path / "board.xlsx"
    _write_board_inventory(xlsx)
    projects = load_projects_from_inventory(xlsx)
    config = AppConfig(projects=projects, runtime=RuntimeConfig(llm_confidence_threshold=0.65))
    record = FileRecord(
        path=str(tmp_path / "202301_CV950D4-B42-12 Specification_v1.0.pdf"),
        filename="202301_CV950D4-B42-12 Specification_v1.0.pdf",
        parent_path=str(tmp_path),
        extension=".pdf",
        size=1,
        mtime=0.0,
    )
    collect_signals(record, config)
    apply_rules(record, config)
    assert record.project_id == projects[0].id
    assert record.classified_by == "rule"
    assert "202301" in record.reason


def test_llm_candidate_json_includes_inventory_fields(tmp_path: Path) -> None:
    project = ProjectDef(
        id="cv950",
        name="CV950D4基于Android R的欧规全制式智能数字电视系统软件",
        inventory_code="202301",
        solution_name="CV950D4",
        board_type="CV950D4-B42",
        year="2023",
        codes=["202301"],
    )
    payload = _compact_candidates([project])[0]
    assert payload["inventory_code"] == "202301"
    assert payload["solution_name"] == "CV950D4"
    assert payload["board_type"] == "CV950D4-B42"
    assert payload["year"] == "2023"


def test_resolve_inventory_in_source_dir(tmp_path: Path) -> None:
    config_path = tmp_path / "projects.yaml"
    config_path.write_text("paths:\n  inventory_excel: list.xlsx\n", encoding="utf-8")
    source = tmp_path / "待整理"
    source.mkdir()
    inv = source / "list.xlsx"
    _write_inventory(inv)
    config = AppConfig(paths=PathsConfig(inventory_excel="list.xlsx"))
    resolved = resolve_inventory_path(config_path, config, source)
    assert resolved == inv.resolve()


def test_resolve_inventory_recursive_in_source(tmp_path: Path) -> None:
    config_path = tmp_path / "projects.yaml"
    source = tmp_path / "待整理"
    nested = source / "2026" / "清单"
    nested.mkdir(parents=True)
    inv = nested / "国高资料清单2026 版型清单.xlsx"
    _write_inventory(inv)
    config = AppConfig(paths=PathsConfig(inventory_excel="国高资料清单2026 版型清单.xlsx"))
    resolved = resolve_inventory_path(config_path, config, source)
    assert resolved == inv.resolve()


def test_resolve_inventory_prefers_shallowest_match(tmp_path: Path) -> None:
    config_path = tmp_path / "projects.yaml"
    source = tmp_path / "待整理"
    shallow = source / "list.xlsx"
    deep = source / "a" / "b" / "list.xlsx"
    deep.parent.mkdir(parents=True)
    _write_inventory(shallow)
    _write_inventory(deep)
    config = AppConfig(paths=PathsConfig(inventory_excel="list.xlsx"))
    resolved = resolve_inventory_path(config_path, config, source)
    assert resolved == shallow.resolve()


def test_scan_excludes_inventory_only_when_filtered_in_pipeline(tmp_path: Path) -> None:
    inbox = tmp_path / "in"
    inbox.mkdir()
    inv = inbox / "list.xlsx"
    _write_inventory(inv)
    (inbox / "202432--CSP311测试报告.pdf").write_bytes(b"pdf")

    all_records = scan_directory(inbox, max_files=100)
    assert len(all_records) == 2

    projects = load_projects_from_inventory(inv)
    inv_resolved = inv.resolve()
    filtered = [r for r in all_records if Path(r.path).resolve() != inv_resolved]
    assert len(filtered) == 1
    assert filtered[0].filename.endswith(".pdf")


def test_inventory_missing_raises(tmp_path: Path) -> None:
    config_path = tmp_path / "projects.yaml"
    config_path.touch()
    config = AppConfig(paths=PathsConfig(inventory_excel="missing.xlsx"))
    source = tmp_path / "in"
    source.mkdir()
    with pytest.raises(Exception, match="未找到项目清单"):
        resolve_inventory_path(config_path, config, source)


def test_empty_inventory_workbook_raises(tmp_path: Path) -> None:
    xlsx = tmp_path / "empty.xlsx"
    Workbook().save(xlsx)
    with pytest.raises(InventoryError, match="未解析到任何项目"):
        load_projects_from_inventory(xlsx)
